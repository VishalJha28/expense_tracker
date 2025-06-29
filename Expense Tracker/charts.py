import tkinter as tk
from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
import os

from utils import raise_screen

def build_chart_screen(chart_screen, root, screens, file_path):
    for widget in chart_screen.winfo_children():
        widget.destroy()

    tk.Label(chart_screen, text="Expense Charts", font=("Segoe UI", 22, "bold"), bg="#f0f2f5").pack(pady=10)

    if not os.path.exists(file_path):
        tk.Label(chart_screen, text="No data available.", font=("Segoe UI", 12), bg="#f0f2f5").pack()
        return

    # Load workbook and filter rows
    wb = load_workbook(file_path)
    ws = wb.active

    rows = [
        row for row in ws.iter_rows(min_row=2, values_only=True)
        if row and all(cell is not None and str(cell).strip() != '' for cell in row[:5])
    ]

    if not rows:
        tk.Label(chart_screen, text="No valid data to plot.", font=("Segoe UI", 12), bg="#f0f2f5").pack()
        return

    totals_by_category = defaultdict(float)
    totals_by_day = defaultdict(float)
    for row in rows:
        try:
            date_str, _, amount, _, category = row[:5]
            date = datetime.strptime(date_str.strip().lstrip("'"), "%Y-%m-%d")
            amount = float(amount)
            totals_by_category[category] += amount
            totals_by_day[date.strftime("%Y-%m-%d")] += amount
        except Exception as e:
            print(f"Skipping row {row}: {e}")
            continue

    # === Bar Chart: Expenses by Category ===
    fig1, ax1 = plt.subplots(figsize=(10, 6), dpi=100)
    categories = list(totals_by_category.keys())
    amounts = list(totals_by_category.values())

    ax1.bar(categories, amounts, color="#4caf50")
    ax1.set_title("Expenses by Category", fontsize=18)
    ax1.set_ylabel("Amount (₹)", fontsize=14)
    ax1.tick_params(axis='x', rotation=45, labelsize=12)
    fig1.tight_layout()

    canvas1 = FigureCanvasTkAgg(fig1, master=chart_screen)
    canvas1.draw()
    canvas1.get_tk_widget().pack(padx=20, pady=(10, 15), fill="both", expand=True)

    # === Line Chart: Daily Expenses ===
    fig2, ax2 = plt.subplots(figsize=(10, 6), dpi=100)
    dates = list(totals_by_day.keys())
    values = list(totals_by_day.values())

    ax2.plot(dates, values, marker="o", linestyle="-", color="#2196f3")
    ax2.set_title("Daily Expenses", fontsize=18)
    ax2.set_ylabel("Amount (₹)", fontsize=14)
    ax2.tick_params(axis='x', rotation=45, labelsize=10)
    fig2.tight_layout()

    canvas2 = FigureCanvasTkAgg(fig2, master=chart_screen)
    canvas2.draw()
    canvas2.get_tk_widget().pack(padx=20, pady=(0, 20), fill="both", expand=True)

    # Back Button
    tk.Button(chart_screen, text="Back", command=lambda: raise_screen(screens['main']),
              bg="#4caf50", fg="white", font=("Segoe UI", 14), relief="flat",
              activebackground="#45a049", padx=20, pady=10).pack(pady=(0, 20))
