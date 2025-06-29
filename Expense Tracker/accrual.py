import tkinter as tk
from tkinter import ttk
from tkcalendar import Calendar
from openpyxl import load_workbook
from datetime import datetime, timedelta
from collections import defaultdict
import os

from utils import raise_screen

def build_accrual_screen(accrual_screen, root, screens, file_path):
    for widget in accrual_screen.winfo_children():
        widget.destroy()

    tk.Label(accrual_screen, text="Accrual Accounting", font=("Segoe UI", 22, "bold"), bg="#f0f2f5").pack(pady=10)

    if not os.path.exists(file_path):
        tk.Label(accrual_screen, text="No data available.", font=("Segoe UI", 12), bg="#f0f2f5").pack()
        return

    wb = load_workbook(file_path)
    ws = wb.active

    rows = [
        row for row in ws.iter_rows(min_row=2, values_only=True)
        if row and all(cell is not None and str(cell).strip() != '' for cell in row[:7])
    ]

    accruals_by_day = defaultdict(float)
    total_assets = 0
    total_capex = 0
    total_opex = 0
    monthly_opex = 0
    today = datetime.today()
    current_month = today.month
    current_year = today.year

    for row in rows:
        try:
            date_str, _, amount, _, _, useful_life, is_asset = row[:7]
            amount = float(amount)
            useful_life = int(useful_life)
            is_asset = str(is_asset).strip().lower() == 'yes'
            date = datetime.strptime(date_str, "%Y-%m-%d")

            if is_asset:
                total_assets += amount
                continue

            if useful_life > 1:
                total_capex += amount
            else:
                if date.month == current_month and date.year == current_year:
                    monthly_opex += amount
                total_opex += amount

            # Accrue daily over the life
            per_day = amount / useful_life
            for i in range(useful_life):
                accrual_day = date + timedelta(days=i)
                if accrual_day.date() > datetime.today().date():
                    break
                accruals_by_day[accrual_day.strftime("%Y-%m-%d")] += per_day
        except Exception as e:
            print(f"Skipping row {row}: {e}")
            continue

    # Summary box
    summary_frame = tk.LabelFrame(accrual_screen, text="Summary", font=("Segoe UI", 12, "bold"), bg="#f0f2f5")
    summary_frame.pack(fill="x", padx=20, pady=10)

    tk.Label(summary_frame, text=f"   Total Assets: ₹{round(total_assets, 2)}", font=("Segoe UI", 11), bg="#f0f2f5").pack(anchor="w")
    tk.Label(summary_frame, text=f"   CapEx: ₹{round(total_capex, 2)}", font=("Segoe UI", 11), bg="#f0f2f5").pack(anchor="w")
    tk.Label(

    summary_frame,

    text=f"   OpEx: ₹{round(total_opex, 2)} (this month: ₹{round(monthly_opex, 2)})",

    font=("Segoe UI", 11),

    bg="#f0f2f5"

).pack(anchor="w")


    # Calendar + Day Breakdown
    calendar_frame = tk.LabelFrame(accrual_screen, text="View Accruals by Day", font=("Segoe UI", 12, "bold"), bg="#f0f2f5")
    calendar_frame.pack(fill="x", padx=20, pady=10)

    cal = Calendar(calendar_frame, selectmode='day', date_pattern='yyyy-mm-dd')
    cal.pack(pady=10)

    result_var = tk.StringVar()
    tk.Label(calendar_frame, textvariable=result_var, font=("Segoe UI", 12, "bold"), bg="#f0f2f5").pack(pady=5)

    def update_selected_day(event=None):

        selected = cal.get_date()

        result_var.set(f"Accrued on {selected}: ₹{round(accruals_by_day.get(selected, 0), 2)}")

    

    cal.bind("<<CalendarSelected>>", update_selected_day)  # Trigger on date pick

    update_selected_day()  # Also run immediately for today's date

    
    # Monthly + Yearly totals
    totals_by_month = defaultdict(float)
    totals_by_year = defaultdict(float)

    for day, amount in accruals_by_day.items():
        dt = datetime.strptime(day, "%Y-%m-%d")
        month_key = dt.strftime("%B %Y")
        year_key = dt.strftime("%Y")
        totals_by_month[month_key] += amount
        totals_by_year[year_key] += amount

    def show_totals(title, totals_dict):
        frame = tk.LabelFrame(accrual_screen, text=title, font=("Segoe UI", 12, "bold"), bg="#f0f2f5")
        frame.pack(fill="x", padx=20, pady=8)
        for key, total in sorted(totals_dict.items()):
            tk.Label(frame, text=f"   {key}: ₹{round(total, 2)}", font=("Segoe UI", 11), bg="#f0f2f5").pack(anchor="w")

    sorted_months = sorted(

    totals_by_month.items(),

    key=lambda x: datetime.strptime(x[0], "%B %Y")

)
    show_totals("Accrued by Month", dict(sorted_months))
    show_totals("Accrued by Year", totals_by_year)

    # Back button
    tk.Button(accrual_screen, text="Back", command=lambda: raise_screen(screens['main']),
              bg="#4caf50", fg="white", font=("Segoe UI", 14), relief="flat",
              activebackground="#45a049", padx=20, pady=10).pack(pady=(0, 20))
