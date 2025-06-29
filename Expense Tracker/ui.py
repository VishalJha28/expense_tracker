import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

from utils import raise_screen
from views import build_view_screen
from charts import build_chart_screen
from accrual import build_accrual_screen

def create_input_field(parent, label_text, var):
    frame = tk.Frame(parent, bg="#f0f2f5")
    frame.pack(fill="x", pady=5)
    tk.Label(frame, text=label_text, font=("Segoe UI", 12), bg="#f0f2f5").pack(anchor="w")
    entry = tk.Entry(frame, textvariable=var, font=("Segoe UI", 14), bg="#ffffff",
                     relief="flat", highlightthickness=1, highlightbackground="#ccc")
    entry.pack(fill="x", ipady=10, pady=5)
    return entry

def create_dropdown_field(parent, label_text, var, options):
    frame = tk.Frame(parent, bg="#f0f2f5")
    frame.pack(fill="x", pady=5)
    tk.Label(frame, text=label_text, font=("Segoe UI", 12), bg="#f0f2f5").pack(anchor="w")
    dropdown = ttk.Combobox(
        frame,
        textvariable=var,
        values=options,
        font=("Segoe UI", 14),
        state="readonly"
    )
    dropdown.current(0)
    dropdown.pack(fill="x", ipady=10, pady=5)
    return dropdown

def create_checkbox(parent, label_text, var):
    frame = tk.Frame(parent, bg="#f0f2f5")
    frame.pack(fill="x", pady=5)
    checkbox = tk.Checkbutton(frame, text=label_text, variable=var, bg="#f0f2f5", font=("Segoe UI", 12))
    checkbox.pack(anchor="w", pady=5)
    return checkbox

def build_main_screen(main_screen, root, screens, file_path):
    category_options = [
        "Select Category", "Food", "Commute", "Q-commerce", "Rent", "Furniture Rent",
        "Electricity", "Gas", "Entertainment", "Shopping", "Health", "Travel", "Other"
    ]
    standard_useful_life = {
      "Rent": "30",
      "Furniture Rent": "30",
      "Gas": "60",
      "Electricity": "30",
      "Internet/mobile": "30"
    }

    vars_dict = {
        'amount': tk.StringVar(),
        'description': tk.StringVar(),
        'category': tk.StringVar(),
        'useful_life': tk.StringVar(),
        'is_asset': tk.BooleanVar()
    }

    def save_expense():
        amount = vars_dict['amount'].get().strip()
        description = vars_dict['description'].get().strip()
        category = vars_dict['category'].get().strip()
        useful_life = vars_dict['useful_life'].get() or standard_useful_life.get('category') or "1"
        is_asset = vars_dict['is_asset'].get()

        if not amount or not description or category == category_options[0]:
            messagebox.showwarning("Missing Info", "Please fill all fields.")
            return

        try:
            float(amount)
        except:
            messagebox.showerror("Invalid", "Amount must be a number.")
            return

        now = datetime.now()
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Date", "Time", "Amount", "Description", "Category", "Useful Life", "Is Asset"])

        ws.append([
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S"),
            amount,
            description,
            category,
            useful_life,
            "Yes" if is_asset else "No"
        ])
        wb.save(file_path)

        messagebox.showinfo("Success", "Expense saved!")
        for var in vars_dict.values():
            var.set("")
        vars_dict['is_asset'].set(False)
        entry_amount.focus_set()

    def open_view_screen():
        build_view_screen(screens['view'], root, screens, file_path)
        raise_screen(screens['view'])

    def open_chart_screen():
        build_chart_screen(screens['charts'], root, screens, file_path)
        raise_screen(screens['charts'])

    center_frame = tk.Frame(main_screen, bg="#f0f2f5")
    center_frame.pack(expand=True)

    tk.Label(center_frame, text="Expense Tracker", font=("Segoe UI", 26, "bold"), bg="#f0f2f5").pack(pady=(10, 20))

    form_frame = tk.Frame(center_frame, bg="#f0f2f5")
    form_frame.pack(fill="x", padx=30, pady=(0, 20))

    entry_amount = create_input_field(form_frame, "Amount", vars_dict['amount'])
    create_input_field(form_frame, "Description", vars_dict['description'])
    create_dropdown_field(form_frame, "Category", vars_dict['category'], category_options)
    create_input_field(form_frame, "Useful Life (Days)", vars_dict['useful_life'])
    create_checkbox(form_frame, "Mark as Asset", vars_dict['is_asset'])

    btn_frame = tk.Frame(center_frame, bg="#f0f2f5")
    btn_frame.pack(pady=(0, 30))

    tk.Button(btn_frame, text="Save Expense", command=save_expense,
              bg="#4caf50", fg="white", font=("Segoe UI", 14), relief="flat",
              width=20, height=2).pack(pady=(0, 10))

    tk.Button(btn_frame, text="View Expenses", command=open_view_screen,
              bg="#2196f3", fg="white", font=("Segoe UI", 14), relief="flat",
              width=20, height=2).pack(pady=(0, 10))

    tk.Button(btn_frame, text="View Charts", command=open_chart_screen, bg="#ff9800", fg="white", font=("Segoe UI", 14), relief="flat",
              width=20, height=2).pack()
    tk.Button(btn_frame, text="View Accruals",

          command=lambda: [build_accrual_screen(screens['accrual'], root, screens, file_path), raise_screen(screens['accrual'])],

          bg="#9c27b0", fg="white", font=("Segoe UI", 14), relief="flat",

          width=20, height=2).pack(pady=(0, 10))



    return entry_amount
