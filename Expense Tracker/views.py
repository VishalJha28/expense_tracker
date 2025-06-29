# views.py

import tkinter as tk
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime
from utils import raise_screen
import os

def build_view_screen(frame, root, screens, file_path):
    def view_expenses():
        for widget in frame.winfo_children():
            widget.destroy()

        tk.Label(frame, text="Expense Details", font=("Segoe UI", 22, "bold"), bg="#f0f2f5").pack(pady=10)

        if not os.path.exists(file_path):
            tk.Label(frame, text="No data yet.", font=("Segoe UI", 12), bg="#f0f2f5").pack()
            return

        wb = load_workbook(file_path)
        ws = wb.active
        rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if row and any(cell is not None for cell in row[:5])]

        table_frame = tk.LabelFrame(frame, text="Recent Entries", bg="#f0f2f5", font=("Segoe UI", 14, "bold"))
        table_frame.pack(fill="x", padx=20, pady=20)

        table = tk.Text(table_frame, font=("Segoe UI", 10), height=6, bg="#ffffff", relief="flat")
        for row in rows[-6:]:
            table.insert(tk.END, f"{row[0]} | ₹{row[2]} | {row[3]} | \n")
        table.config(state="disabled")
        table.pack(fill="x")

        # Totals
        #all_rows = list(ws.iter_rows(min_row=2, values_only=True))
        totals_by_category = defaultdict(float)
        totals_by_day = defaultdict(float)
        totals_by_week = defaultdict(float)
        totals_by_month = defaultdict(float)

        for row in rows:
            if row[0] is None:
                break
            date_str, _, amount, _, category = row[:5]
            date = datetime.strptime(date_str.strip().lstrip("'"), "%Y-%m-%d")
            amount = float(amount)
            totals_by_category[category] += amount
            totals_by_day[date.strftime("%Y-%m-%d")] += amount
            totals_by_week[date.strftime("Week %U, %Y")] += amount
            totals_by_month[date.strftime("%B %Y")] += amount

        def show_totals(title, totals_dict):
            key = lambda kv: -kv[1] if title == "Total by Category" else kv
            section = tk.LabelFrame(frame, text=title, font=("Segoe UI", 12, "bold"), bg="#f0f2f5")
            section.pack(fill="x", padx=20, pady=8)
            for key, total in sorted(totals_dict.items(),key=key)[-5:]:
                tk.Label(section, text=f"   {key}: ₹{round(total, 2)}", font=("Segoe UI", 11), bg="#f0f2f5").pack(anchor="w")

        show_totals("Total by Category", totals_by_category)
        show_totals("Total by Day", totals_by_day)
        show_totals("Total by Week", totals_by_week)
        show_totals("Total by Month", totals_by_month)

        tk.Button(frame, text="Back", command=lambda: raise_screen(screens['main']),
                  bg="#4caf50", fg="white", font=("Segoe UI", 14), relief="flat",
                  activebackground="#45a049", padx=20, pady=10).pack(pady=20)

    # Call once to render content
    view_expenses()
